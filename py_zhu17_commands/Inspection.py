import openpyxl      #xlrd更新到了2.0.1版本，只支持.xls文件，不支持.xlsx
import time
from netmiko import Netmiko
import os,sys
import shutil

os.chdir(sys.path[0]) 

date = time.strftime("%Y%m%d", time.localtime())

def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)


def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]

    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组
    


def show_to_txt(x_namepass,x_command):
    b = len(x_namepass)
    d = len(x_command)

    newfile = ('./'+date)
    if not os.path.exists(newfile):       #判断是否有相同名字的文件夹，有的话覆盖掉
        os.makedirs(newfile)
        os.chdir(newfile)
    else:
        shutil.rmtree(newfile)           #removes all the subdirectories!
        os.makedirs(newfile)
        os.chdir(newfile)


    for j in range(1,b):
        print('Now is checking '+x_namepass[j][1]+',the ip address is '+x_namepass[j][2])
        my_device = {
            "host": x_namepass[j][2],
            "username": x_namepass[j][3],
            "password": x_namepass[j][4],
            "device_type": "cisco_ios",
        }
        
        net_connect = Netmiko(**my_device)
        
        write_file = open(x_namepass[j][1]+'_'+date+'.txt', 'w')   #模式为追加
        for n in range(0,d):
            print(x_command[n][1], file=write_file)     #往write_file里面写入
            print("\n", file=write_file)
            
            output = net_connect.send_command(x_command[n][1])
            
            print(output, file=write_file)
            print("\n=============================================\n", file=write_file)
        write_file.close()
        
        net_connect.disconnect()
    
    os.chdir('..')






if __name__ == '__main__':
    t_namepass = excel_to_list2d('Zhu_list.xlsx', 0)
    t_command = excel_to_list2d('Zhu_list.xlsx', 1)

    show_to_txt(t_namepass,t_command)








