import os,sys
from netmiko import Netmiko,ConnectHandler
import openpyxl

os.chdir(sys.path[0]) 


ip = '10.16.63.103'
username = "admin"
password = "admin@123"
#show_int = "show interface E2/47"



def get_data():

    my_device = {
    "host": ip,
    "username": username,
    "password": password,
    "device_type": "cisco_ios",
    }

    try:
        net_connect = ConnectHandler(**my_device)        
        tdata_original = net_connect.send_command('sh int | in ernet|CRC')
        net_connect.disconnect()

        return tdata_original

    except:
        print('Something wrong')

    


def CRC_to_list(xdata_original):
#将文件逐行读取到一个列表且忽略空行
    tall_lines = xdata_original.splitlines()
    tlist1 = [line.strip() for line in tall_lines if line.strip()]

    tlist2 = []
    tlist3 = []
    for tline in tlist1:
        if ' is ' in tline:
            tlist2.append(tline.split( )[0])
        if 'CRC' in tline:
            tlist3.append(tline.split( )[4])
        else:
            pass

    #print(tlist2)
    #print(tlist3)
    return tlist2,tlist3

def list_to_xlsx(xlist1,xlist2):        #把list写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist1)):    #两个列表是一对一关系，所以只用一个i
        #ws.append(xlist1[i])    #append的对象只能是列表，元组什么的
        ws.cell(row = i+1, column = 1).value = xlist1[i]
        ws.cell(row = i+1, column = 2).value = xlist2[i]   # 写入位置的行列号可以任意改变

    wb.save('CRC.xlsx')

if __name__ == '__main__':
    mtdata_original = get_data()
    mlist2,mlist3 = CRC_to_list(mtdata_original)
    #print(mlist2[0])
    #print(mlist3[0])
    list_to_xlsx(mlist2,mlist3)
    


