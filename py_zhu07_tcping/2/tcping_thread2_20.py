import os,sys
import openpyxl
from netmiko import Netmiko
import threading
import queue
import subprocess

os.chdir(sys.path[0]) 

q = queue.Queue()


def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)

def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]
    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组





def tcp_ping():
#导入要执行的命令行列表

    while not q.empty():
        tcping_one = q.get() 
        
        output = subprocess.getoutput(tcping_one)     #逐行执行命令行，并获取执行结果

        lock.acquire()   #同步锁
        print('Checking %s ......' %tcping_one)
        #print(output)          #把执行结果输出在屏幕上,会不按顺序来，无法避免
        if 'open' in output:            #如果输出结果中有（time）out字样，tcping失败
            result_d[tcping_one] = 'ok'
        else:
            result_d[tcping_one] = 'ng'

        lock.release()   #同步锁





if __name__ == '__main__':

    result_d = {}    #定义一个用于存放结果的空字典

    tcping_list2d = excel_to_list2d('TCPING List.xlsx', 0)     #得到的是二维列表
    tcping_list = sum(tcping_list2d,[])        #变成一维列表

    for i in range(len(tcping_list)):
        q.put(tcping_list[i])           #把tcping命令放入线程队列
    
    WORD_THREAD = 20       # 定义工作线程    
    threads = []
    
    lock = threading.Lock()

    for i in range(WORD_THREAD):
        thread = threading.Thread(target=tcp_ping)
        thread.start()
        threads.append(thread)
    for thread in threads:
        thread.join()

    #print(result_d)


    wb = openpyxl.Workbook()
    ws = wb.active


    for i in range(len(tcping_list)):       
        tcping_list2d[i].append(result_d[tcping_list[i]])    #在每个二维列表的值后面把字典里key对应的值接上

        ws.append(tcping_list2d[i])
    wb.save('TCPING Result.xlsx')

    print('All finished, pls check "TCPING Result.xlsx"')

