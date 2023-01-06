import os,sys
import openpyxl
import threading
import queue
import numpy
import subprocess

os.chdir(sys.path[0]) 

q = queue.Queue()

def Open_Excel(xfile):
#打开要解析的Excel文件
    try:
        t_excel_data = openpyxl.load_workbook(xfile)
        t_sheet_names = t_excel_data.sheetnames
        return t_excel_data,t_sheet_names    #所有sheet名字的列表
    except Exception as e:
      print(e)



def Ping_One_IP():
#从线程列表中拿ip
    while not q.empty():
        ip = q.get()  
        
        row = [] 
        row.append(ip)

        if 'TTL' in subprocess.getoutput("ping "+ip):
            shifou = 'ok'
        else:
            shifou = 'ng'
        row.append(shifou)

        temp_result_list2d.append(row)


def List2d_To_Dict(xlist2d):
#把列表中的第一个元素作为字典的key，剩下的作为值
    xdict = {}
    for i in range(len(xlist2d)):
        key = xlist2d[i][0]         
        tlist = xlist2d[i]
        tlist.pop(0)    #删掉第一个元素
        value = tlist
        xdict[key] = value
    return(xdict)






def add_list2d_to_xlsx(x_list2d,x_sheet_name):        #把新的list追加到现有的Excel文件
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('Result.xlsx')
    ws = wb.create_sheet(x_sheet_name)
    # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
    for i in range(len(x_list2d)):
        for j in range(len(x_list2d[i])):
            ws.cell(row = i+1, column = j+1).value = x_list2d[i][j]
    # 保存操作
    wb.save('Result.xlsx')




if __name__ == '__main__':

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save('Result.xlsx')

    m_excel_data,m_sheet_names = Open_Excel('IP addresses.xlsx')

    for m_sheet_name in m_sheet_names:
        print('Checking %s sheet......' %m_sheet_name) 

        m_sheet = m_excel_data[m_sheet_name]
        ip_list = []
        for row in range(2,m_sheet.max_row+1):
            ip_list.append(m_sheet.cell(row,1).value)
        #print(ip_list)

        temp_result_list2d = []     #全局数组，多线程的结果汇集

        for i in range(len(ip_list)):
            q.put(ip_list[i])           #把ip地址放入线程队列

        THREAD = 20           #线程数
        threads = []

        for i in range(THREAD):       
            thread = threading.Thread(target = Ping_One_IP)
            thread.start()
            threads.append(thread)
        for thread in threads:
            thread.join()

        temp_dict = {}
        temp_dict = List2d_To_Dict(temp_result_list2d)

        one_sheet_result2d = []
        title_list = []
        title_list.append('IP')
        title_list.append('Result')
        one_sheet_result2d.append(title_list)
        ip_list2d = numpy.array(ip_list).reshape(-1,1)    #把数据类型转换为数组，再reshape成二维数组，直接把原来读取的IP列表进行转换
        for i in range(len(ip_list)):
            row = ip_list2d[i].tolist()    #把数据类型重新转换为列表
            row.extend(temp_dict[ip_list[i]])    #把字典里key对应的值重新接上
            #print(row)
            one_sheet_result2d.append(row)

        
        add_list2d_to_xlsx(one_sheet_result2d,m_sheet_name)

    wb = openpyxl.load_workbook('Result.xlsx')
    del wb['Sheet']         #删掉一开始生成的默认sheet
    wb.save('Result.xlsx')

        
