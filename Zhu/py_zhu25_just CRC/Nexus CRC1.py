import os,sys
from netmiko import Netmiko,ConnectHandler
import openpyxl

os.chdir(sys.path[0]) 


username = "admin"
password = "admin@123"



def xlsx_to_2dlist():        
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('Hostname IP.xlsx')
    ws = wb['Sheet1']
 
    i2dlist = []
    for row in ws.rows:
        irow = []
        for cell in row:
            #print(cell.value)
            irow.append(cell.value)
        i2dlist.append(irow)

    # 保存操作
    wb.save('Hostname IP.xlsx')

    return i2dlist






def get_data(xip):

    my_device = {
    "host": xip,
    "username": username,
    "password": password,
    "device_type": "cisco_ios",
    }

    try:
        net_connect = ConnectHandler(**my_device)        
        tdata_original = net_connect.send_command('sh int | in ernet|CRC')
        net_connect.disconnect()

        return tdata_original

    except:
        print('Something wrong')

    


def CRC_to_list(xdata_original):
#将文件逐行读取到一个列表且忽略空行
    tall_lines = xdata_original.splitlines()
    tlist1 = [line.strip() for line in tall_lines if line.strip()]

    tlist2 = []
    tlist3 = []
    for tline in tlist1:
        if ' is ' in tline:
            tlist2.append(tline.split( )[0])
        if 'CRC' in tline:
            tlist3.append(tline.split( )[4])
        else:
            pass

    #print(tlist2)
    #print(tlist3)
    return tlist2,tlist3

def list_to_xlsx(xlist1,xlist2):        #把list写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist1)):    #两个列表是一对一关系，所以只用一个i
        #ws.append(xlist1[i])    #append的对象只能是列表，元组什么的
        ws.cell(row = i+1, column = 1).value = xlist1[i]
        ws.cell(row = i+1, column = 2).value = xlist2[i]   # 写入位置的行列号可以任意改变

    wb.save('CRC.xlsx')

if __name__ == '__main__':
    m2dlist = xlsx_to_2dlist()

    re_ex= openpyxl.Workbook('ERROR check.xlsx')    #新建一个工作簿并命名

    for i in range(len(m2dlist)):
        ws2 = re_ex.create_sheet(title = m2dlist[i][0])        #新建sheet并设定sheet名称
        ws = re_ex.active
        ws.cell(row = 1, column = 2).value = "ri"
        #for j in range(len(m2dlist)):
            #ws2.append(m2dlist[j])

    re_ex.save('ERROR check.xlsx')


    '''
    mtdata_original = get_data()
    mlist2,mlist3 = CRC_to_list(mtdata_original)
    #print(mlist2[0])
    #print(mlist3[0])
    list_to_xlsx(mlist2,mlist3)
    '''
    

'''
def open_excel(xfile):
#打开要解析的Excel文件
    try:
        data = xlrd.open_workbook(xfile)
        return data
    except Exception as e:
      print(e)


def excel_to_list2d(excel_file, by_index):   #by_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    data = open_excel(excel_file)
    table = data.sheets()[by_index]

    totalarray = []
    for row in range(table.nrows):
        subarray = []
        for col in range(table.ncols):
            subarray.append(table.cell(row,col).value)
        totalarray.append(subarray)

    return(totalarray)   #返回这个二维数组



def list2d_to_xlsx(xlist2d,x_excel_name):        #把list写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save(x_excel_name)
'''
