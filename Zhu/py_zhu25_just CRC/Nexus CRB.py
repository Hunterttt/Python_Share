import os,sys
from netmiko import Netmiko,ConnectHandler
import openpyxl

os.chdir(sys.path[0]) 

'''
ip = '10.16.63.103'
username = "admin"
password = "admin@123"
#show_int = "show interface E2/47"
'''

'''
def get_data():

    my_device = {
    "host": ip,
    "username": username,
    "password": password,
    "device_type": "cisco_ios",
    }

    try:
        net_connect = ConnectHandler(**my_device)        
        tdata_original = net_connect.send_command('sh int | in ernet|CRC')
        net_connect.disconnect()

        return tdata_original

    except:
        print('Something wrong')
'''
    


def CRC_to_list(xfile):
#将文件逐行读取到一个列表且忽略空行
    with open(xfile, "r") as tfile:
        #print(type(tfile))
        tlist1 = [line.strip() for line in tfile if line.strip()]

    tlist2 = []
    tlist3 = []
    for tline in tlist1:
        if ' is ' in tline and 'Ethernet' in tline:
            tlist2.append(tline.split( )[0])
        if 'CRC' in tline:
            tlist3.append(tline.split( )[4])
        else:
            pass

    #print(tlist2)
    #print(tlist3)
    return tlist2,tlist3

def list_to_xlsx(xlist1,xlist2):        #把list写入Excel文件
    wb = openpyxl.Workbook()
    #ws = wb.active

    ws = wb.create_sheet('ri')
    

    for i in range(len(xlist1)):    #两个列表是一对一关系，所以只用一个i
        #ws.append(xlist1[i])    #append的对象只能是列表，元组什么的
        
        ws.cell(row = i+1, column = 1).value = xlist1[i]
        ws.cell(row = i+1, column = 2).value = xlist2[i]   # 写入位置的行列号可以任意改变

    #让cell的宽度自适应
    dims = {}    #这个字典中将要储存的是 键：值= cell.column_letter单元格的所在列的位置（字母表示）：cell的长度（宽度）
    for row in ws.rows:
        for cell in row:  #遍历每一个cell
            if cell.value:   #排除空cell
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                #dims.get(cell.column_letter, 0)   用来得到上一个循环中cell的宽度
                #len(str(cell.value)) 这个cell的字符串长度
                #cell.column_letter - 单元格的所在列的位置（字母表示）
    for col, value in dims.items():   #col, value分别是字典的键和值，是经过上面循环得到的每一列的列号和最大宽度，.item相当于遍历字典的每一个值
        ws.column_dimensions[col].width = value+2  #多加两个空格长度更好看


    wb.remove(wb['Sheet']) #删除某个工作表
    wb.save('CRC.xlsx')

if __name__ == '__main__':
    #mtdata_original = get_data()
    mlist2,mlist3 = CRC_to_list('Source data.txt')
    #print(mlist2[0])
    #print(mlist3[0])
    list_to_xlsx(mlist2,mlist3)
    


