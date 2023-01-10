import os,sys
from netmiko import Netmiko,ConnectHandler
import openpyxl

os.chdir(sys.path[0]) 


username = "admin"
password = "admin@123"



def xlsx_to_2dlist():        
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('Hostname IP.xlsx')
    ws = wb['Sheet1']
 
    i2dlist = []
    for row in ws.rows:
        irow = []
        for cell in row:
            #print(cell.value)
            irow.append(cell.value)
        i2dlist.append(irow)

    # 保存操作
    wb.save('Hostname IP.xlsx')

    return i2dlist






def get_data(xip):

    my_device = {
    "host": xip,
    "username": username,
    "password": password,
    "device_type": "cisco_ios",
    }

    try:
        net_connect = ConnectHandler(**my_device)        
        tdata_original = net_connect.send_command('sh int | in ernet|CRC')
        net_connect.disconnect()

        #print(tdata_original)
        return tdata_original

    except:
        print('Something wrong')

    


def CRC_to_list(xdata_original):
#将文件逐行读取到一个列表且忽略空行
    tall_lines = xdata_original.splitlines()
    tlist1 = [line.strip() for line in tall_lines if line.strip()]

    tlist2 = []
    tlist3 = []
    for tline in tlist1:
        if ' is ' in tline and 'Ethernet' in tline:
            tlist2.append(tline.split( )[0])
        if 'CRC' in tline:
            tlist3.append(tline.split( )[4])
        else:
            pass

    #print(tlist2)
    #print(tlist3)
    return tlist2,tlist3

'''
def list_to_xlsx(xlist1,xlist2):        #把list写入Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(xlist1)):    #两个列表是一对一关系，所以只用一个i
        #ws.append(xlist1[i])    #append的对象只能是列表，元组什么的
        ws.cell(row = i+1, column = 1).value = xlist1[i]
        ws.cell(row = i+1, column = 2).value = xlist2[i]   # 写入位置的行列号可以任意改变

    wb.save('CRC.xlsx')
'''


if __name__ == '__main__':
    m2dlist = xlsx_to_2dlist()
    print(m2dlist)
    mbook= openpyxl.Workbook()    #新建一个工作簿，不要起名字，否则下面的cell属性会为无效，但是append能用
    #或者先创建一个xlsx，save之后再重新打开

    for i in range(len(m2dlist)):

        mdata_original = get_data(m2dlist[i][1])

        mlist2,mlist3 = CRC_to_list(mdata_original)

        msheet = mbook.create_sheet(title = m2dlist[i][0])        #新建sheet并设定sheet名称

        for j in range(len(mlist2)):
            msheet.cell(row = j+1, column = 1).value = mlist2[j]
            msheet.cell(row = j+1, column = 2).value = mlist3[j]   # 写入位置的行列号可以任意改变
            
        mbook.remove(mbook['Sheet']) #删除某个工作表

    mbook.save('ERROR check.xlsx')


    '''
    mtdata_original = get_data()
    mlist2,mlist3 = CRC_to_list(mdata_original)
    #print(mlist2[0])
    #print(mlist3[0])
    list_to_xlsx(mlist2,mlist3)
    '''
    


