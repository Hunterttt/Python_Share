import os,sys,time
from netmiko import Netmiko,ConnectHandler
import openpyxl

os.chdir(sys.path[0]) 


ip = '192.168.32.101'
username = "admin"
password = "Admin@123"
show_int = "show int g0/0"


def get_data(ip):

    my_device = {
    "host": ip,
    "username": username,
    "password": password,
    "device_type": "cisco_ios",
    }


    try:
        net_connect = ConnectHandler(**my_device)        
        tline1_original = net_connect.send_command('show clock')
        tline1_list = tline1_original.split("\n")
        tline1 = list(filter(lambda x:len(x)>1,tline1_list))
        tline2 = ['runts','giants','CRC','no buffer','input error','short frame','overrun','underrun','ignored']
        tline3_original = net_connect.send_command(show_int)

        net_connect.disconnect()


    except:
        print('Something wrong')

    return tline1,tline2,tline3_original


def get_numbs(xlist):
    for tline in xlist:
        if 'throttles' in tline:
            tlist1 = tline.split(" ")
            tlist2 = list(filter(lambda x:len(x)>0,tlist1))    #只挑出长度大于0的
            #print(tlist2)
        elif 'ignored' in tline:
            tlist3 = tline.split(" ")
            tlist4 = list(filter(lambda x:len(x)>0,tlist3))
            #print(tlist4)
        else:
            pass

    tlist = tlist2 + tlist4

    num_list = []
    for tline in tlist:
        try:
            num_list.append(int(tline))       #挑出里面的int
        except:
            pass
    return num_list







def add_list2d_to_xlsx(xlist2d,x,y):
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('CRC error.xlsx')
    ws = wb['Sheet']
    # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
    for i in range(len(xlist2d)):
        for j in range(len(xlist2d[i])):
            ws.cell(row = x+i, column = y+j).value = xlist2d[i][j]
    # 保存操作
    wb.save('CRC error.xlsx')



if __name__ == '__main__':
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save('CRC error.xlsx')

    i = 1
    while 1:
        
        line1_list,line2_list,line3_data1 = get_data(ip)
        print(line1_list)

        line3_data2 = line3_data1.split("\n")        
        line3_list = get_numbs(line3_data2)
        line_2dlist = []
        line_2dlist.append(line1_list)
        line_2dlist.append(line2_list)
        line_2dlist.append(line3_list)
        add_list2d_to_xlsx(line_2dlist,i,1)
        i = i+4
        time.sleep(3)









    
