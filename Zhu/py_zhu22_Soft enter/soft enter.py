import os,sys
import openpyxl
from openpyxl import Workbook

os.chdir(sys.path[0]) 

def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)

def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet，0表示第一个sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]
    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        #print(subarray[0])
        subarray[0] = subarray[0].replace("&", "\n")    #列表中的元素修改后重新赋值回去
        #print(subarray[0])
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组


def list1d_to_excel(xlist2d):
#写入Excel文件
    wb = Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save('enter.xlsx')


if __name__ == '__main__':
    t_list2d = excel_to_list2d("sample.xlsx",0)
    list1d_to_excel(t_list2d)
