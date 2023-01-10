import time
import os,sys
from netmiko import Netmiko
import openpyxl

os.chdir(sys.path[0]) 

ip1 = "10.80.254.1"


def show_cpu(x_ip):

    my_device = {
    "host": x_ip,
    "username": "admin",
    "password": "Yingke@2021.",
    "device_type": "cisco_ios",
    }
    try:
        net_connect = Netmiko(**my_device)
        cpu_res_all = net_connect.send_command('show processes cpu')
        net_connect.disconnect()

        cpu_res_list = cpu_res_all.splitlines()
        #print(cpu_res_list)
        for i in range(len(cpu_res_list)):        
            if "CPU utilization" in cpu_res_list[i]:
                print(cpu_res_list[i])
                return cpu_res_list[i]
            else:
                pass       
            
    except:
        print("Something wrong")







def add_res_to_xlsx(x_cpu_res,a):
    # 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook('show_cpu.xlsx')
    ws = wb['Sheet']
    ws.cell(row = a-2, column = 1).value = x_cpu_res
    # 保存操作
    wb.save('show_cpu.xlsx')


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save('show_cpu.xlsx')

    i = 1

    while 1:
        t_res = show_cpu(ip1)
        i += 2
        time.sleep(300)
        add_res_to_xlsx(t_res,i)


