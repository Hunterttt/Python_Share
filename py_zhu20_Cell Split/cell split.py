import os,sys
import openpyxl
import numpy
from openpyxl import Workbook

os.chdir(sys.path[0]) 

def open_excel(xfile):
#打开要解析的Excel文件
    try:
        excel_data = openpyxl.load_workbook(xfile)
        return excel_data
    except Exception as e:
      print(e)

def excel_to_list2d(x_excel_file, x_index):   #x_index表示选择excel里面的第几个Sheet
#将excel表中的各个值读入一个二维数组
    t_excel_data = open_excel(x_excel_file)
    t_sheet = t_excel_data.worksheets[x_index]
    totalarray = []
    for row in range(1,t_sheet.max_row+1):
        subarray = []
        for col in range(1,t_sheet.max_column+1):
            subarray.append(t_sheet.cell(row,col).value)
        totalarray.append(subarray)
    #print(totalarray)
    return(totalarray)   #返回这个二维数组


def list2d_to_xlsx(xlist2d):
#写入Excel文件
    wb = Workbook()
    ws = wb.active
    for i in range(len(xlist2d)):
        ws.append(xlist2d[i])
    wb.save('Split Result.xlsx')


if __name__ == '__main__':
    r_cell_list2d = []
    cell_list2d = excel_to_list2d('Need Split.xlsx', 0)
    for i in range(len(cell_list2d)):
        a = 0   #是一个判断量

        for j in range(len(cell_list2d[i])):         #遍历所有单元格
            #print(cell_list2d[i][j])
            if '-' in str(cell_list2d[i][j]):       #注意这个地方要把单元格里的数据统一改成字符串str
                cell_list = str(cell_list2d[i][j]).split('-')     #拆成前后两个数字
                spl = []
                k = 0
                while int(cell_list[0]) + k <= int(cell_list[1]):  #需要计算的地方，要把单元格里的数据改成整型int
                    spl.append(int(cell_list[0])+k)          #将两个数字中的所有数字放在一个列表里
                    k+=1
                #print(spl)
                #print(cell_list2d[i][j-1])
                spl_2d = numpy.array(spl).reshape(-1,1)     #变成二维列表

                for x in spl_2d:
                    
                    y = x.tolist()
                    y.insert(0,cell_list2d[i][j-1])    #在每项前面加ip地址
                    #print(y)
                    r_cell_list2d.append(y)

                a += 1
            else:
                pass
        
        #print(a)
        if a == 0:
            r_cell_list2d.append(cell_list2d[i])    #如果一行里面没有拆分，就直接append那一行
        else:
            pass


    #print(r_cell_list2d)
    list2d_to_xlsx(r_cell_list2d)